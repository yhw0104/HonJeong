# -*- coding: utf-8 -*-
"""혼정 기능요구사항 목록 → Excel 생성 (02-요구사항명세서 + 06-화면별-기능요구사항 기반)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["요구사항ID", "대분류", "중분류", "소분류", "요구사항 내용", "중요도", "상태", "버전", "비고"]

# (ID, 대분류, 중분류, 소분류, 요구사항 내용, 중요도, 상태, 버전, 비고)
ROWS = [
    # ───────── 인증 · 계정 ─────────
    ("AUTH-001", "인증·계정", "소셜 로그인", "카카오 로그인",
     "카카오 OAuth로 회원가입/로그인. 클라가 받은 idToken/authCode를 서버가 카카오에 검증 후 (provider, providerUserId)로 회원 조회/생성, 자체 JWT 발급",
     "높음", "예정", "1.0", "FR-004 / POST /api/auth/oauth/kakao"),
    ("AUTH-002", "인증·계정", "소셜 로그인", "애플 로그인",
     "애플 OAuth로 회원가입/로그인. 애플 토큰 서버 검증 후 회원 조회/생성, 자체 JWT 발급",
     "높음", "예정", "1.0", "FR-004 / POST /api/auth/oauth/apple"),
    ("AUTH-003", "인증·계정", "소셜 로그인", "소셜 계정 연동 매핑",
     "social_accounts에 (provider, provider_user_id) UNIQUE 매핑. 1회원 N소셜(카카오+애플 동시 연동). 공급자 access/refresh 토큰은 저장하지 않음",
     "높음", "예정", "1.0", "ERD A-2"),
    ("AUTH-004", "인증·계정", "휴대폰 인증", "인증번호 발송",
     "휴대폰 번호로 SMS 인증번호 발송. 발송 rate-limit 적용",
     "높음", "예정", "1.0", "FR-005 / POST /api/auth/phone/send-code"),
    ("AUTH-005", "인증·계정", "휴대폰 인증", "인증번호 확인",
     "인증번호 검증(만료 3분·시도횟수 제한). 기존회원=JWT / 신규=온보딩용 임시 토큰 발급",
     "높음", "예정", "1.0", "FR-005 / POST /api/auth/phone/verify"),
    ("AUTH-006", "인증·계정", "약관", "약관 동의",
     "서비스 이용약관·개인정보 처리방침·위치기반(필수 3종) + 마케팅 수신(선택) 동의 저장. 사용자당 1행",
     "높음", "예정", "1.0", "FR-006 / POST /api/auth/terms"),
    ("AUTH-007", "인증·계정", "가입 흐름", "신규/기존 회원 분기",
     "휴대폰 인증 성공 시 신규=ProfileSetup(임시토큰)으로, 기존=바로 메인(JWT)으로 분기",
     "높음", "예정", "1.0", "06 A-3"),
    ("AUTH-008", "인증·계정", "토큰", "토큰 재발급",
     "Refresh 토큰으로 Access 토큰 재발급",
     "높음", "예정", "1.0", "FR-002 / POST /api/auth/refresh"),
    ("AUTH-009", "인증·계정", "토큰", "로그아웃",
     "토큰 폐기 처리(클라 저장 토큰 제거)",
     "높음", "예정", "1.0", "FR-002"),
    ("AUTH-010", "인증·계정", "프로필 셋업", "닉네임 설정·중복확인",
     "닉네임 입력 및 실시간 중복 확인(UNIQUE)",
     "높음", "예정", "1.0", "FR-003E / GET /api/users/nickname-check"),
    ("AUTH-011", "인증·계정", "프로필 셋업", "프로필 사진 업로드",
     "프로필 이미지 업로드(오브젝트 스토리지)",
     "높음", "예정", "1.0", "POST /api/files"),
    ("AUTH-012", "인증·계정", "프로필 셋업", "동네(현재 위치) 설정",
     "현재 위치 역지오코딩으로 동네(시군구·동) + 중심 좌표 설정",
     "높음", "예정", "1.0", "GET /api/geo/reverse"),
    ("AUTH-013", "인증·계정", "프로필 셋업", "성별/연령대 입력",
     "성별(MALE/FEMALE/NONE)·연령대(10s~50+) 입력",
     "높음", "예정", "1.0", "FR-003E"),
    ("AUTH-014", "인증·계정", "프로필 셋업", "한 줄 소개 입력",
     "introduction(한 줄 소개) 입력",
     "중간", "예정", "1.0", "FR-003E"),
    ("AUTH-015", "인증·계정", "프로필 셋업", "식사 성향 설정",
     "같이 먹을 때 성향 dining_style(TALK=대화/QUIET=조용) 설정",
     "높음", "예정", "1.0", "FR-003E"),
    ("AUTH-016", "인증·계정", "프로필 셋업", "선호 음식 선택",
     "선호 음식 최대 3개 선택(user_food_preferences). 마스터는 GET /api/meta/food-categories",
     "중간", "예정", "2.0", "FR-003E"),
    ("AUTH-017", "인증·계정", "가입 흐름", "가입 완료(프로필 저장)",
     "온보딩 임시 토큰으로 프로필 저장 → 가입 완료 및 JWT 발급",
     "높음", "예정", "1.0", "PATCH /api/users/me"),
    ("AUTH-018", "인증·계정", "프로필", "내 프로필 조회",
     "내 프로필(phone·nickname·사진·동네·성향 등) 조회",
     "높음", "예정", "1.0", "FR-003 / GET /api/users/me"),
    ("AUTH-019", "인증·계정", "프로필", "내 프로필 수정",
     "프로필 필드 수정(닉네임·사진·소개·동네·성향 등)",
     "높음", "예정", "1.0", "FR-003 / PATCH /api/users/me"),
    ("AUTH-020", "인증·계정", "프로필", "같이먹기 수신 허용 토글",
     "allow_meal_request opt-in/거부 토글(수신 거부 시 신청 차단)",
     "높음", "예정", "1.0", "FR-108 / NFR-03"),
    ("AUTH-021", "인증·계정", "계정", "회원 탈퇴",
     "회원 탈퇴 처리(status=WITHDRAWN)",
     "중간", "예정", "2.0", "ERD users.status"),

    # ───────── 장소 · 검색 ─────────
    ("PLACE-001", "장소·검색", "검색", "식당 검색",
     "카카오 로컬 API 프록시로 가게명·지역 검색(placeId·명칭·주소·좌표·카테고리)",
     "높음", "예정", "1.0", "FR-105 / GET /api/places/search"),
    ("PLACE-002", "장소·검색", "검색", "주변 식당 거리순 조회",
     "GPS 기준 주변 식당 후보를 거리순 노출(거리·현재혼밥러·메이트수·별점·1인석 태그 포함)",
     "높음", "예정", "1.0", "FR-105 / GET /api/places/nearby"),
    ("PLACE-003", "장소·검색", "캐싱", "선택 식당 캐싱(upsert)",
     "체크인/리뷰/즐겨찾기 시 external_id로 places upsert(없으면 INSERT, 있으면 재사용)",
     "높음", "예정", "1.0", "ERD B-1"),
    ("PLACE-004", "장소·검색", "상세", "식당 기본 정보 조회",
     "식당 상세 기본 정보 조회",
     "높음", "예정", "1.0", "FR-105 / GET /api/places/{placeId}"),
    ("PLACE-005", "장소·검색", "상세", "식당 메뉴",
     "메뉴명·가격·사진(place_menus). 출처 UGC/ADMIN/API",
     "중간", "예정", "2.0", "FR-105E / G3"),
    ("PLACE-006", "장소·검색", "상세", "식당 사진 갤러리",
     "식당 상세 사진 탭 — review_photos를 place_id로 집계해 노출(별도 place_photos 미사용)",
     "중간", "예정", "2.0", "FR-105E / GET /api/places/{placeId}/photos"),
    ("PLACE-007", "장소·검색", "상세", "편의시설/혼밥 요소 칩",
     "바테이블·1인석·와이파이·콘센트·혼밥환영·장기체류(place_facilities). 가게당 1행, true 칩만 렌더",
     "중간", "예정", "2.0", "FR-105E / G3"),
    ("PLACE-008", "장소·검색", "상세", "영업시간/전화/홈페이지",
     "places 확장 필드(business_hours·phone·homepage_url). 출처 미정(G3)",
     "중간", "예정", "2.0", "FR-105E"),
    ("PLACE-009", "장소·검색", "상세", "혼밥 친화도 점수·요소",
     "reviews.solo_friendly_rating 평균 + review_tags 빈도 집계로 식당 혼밥 친화도 산출",
     "중간", "예정", "2.0", "FR-105E / C4 / GET /api/places/{placeId}/solo-friendliness"),

    # ───────── 체크인 ─────────
    ("CHECKIN-001", "체크인", "체크인", "혼밥 체크인 시작",
     "특정 식당 선택 후 '지금 혼밥 중' 상태 등록(토글 ON). 서비스 데이터 시작점",
     "높음", "예정", "1.0", "FR-101 / POST /api/check-ins"),
    ("CHECKIN-002", "체크인", "체크인", "단일 활성 체크인 보장",
     "사용자당 ACTIVE 체크인 1개만(부분 유니크 인덱스). 중복 시 409",
     "높음", "예정", "1.0", "FR-101 / ERD C8"),
    ("CHECKIN-003", "체크인", "체크인", "체크인 종료",
     "혼밥 종료 처리(토글 OFF / 상태 바 '끝내기')",
     "높음", "예정", "1.0", "FR-102 / PATCH /api/check-ins/{id}/end"),
    ("CHECKIN-004", "체크인", "체크인", "종료 누락 자동 만료(TTL)",
     "종료 누락 시 배치/조회 시점 자동 만료로 '현재 N명' 정합성 유지",
     "높음", "예정", "1.0", "NFR-07"),
    ("CHECKIN-005", "체크인", "체크인", "409 충돌 복구",
     "진입 시 GET /check-ins/me 선동기화, 409면 기존 종료 후 재시도, 같은 장소 재요청은 멱등",
     "높음", "예정", "1.0", "ERD §C-1"),
    ("CHECKIN-006", "체크인", "체크인", "내 현재 체크인 조회",
     "현재 ACTIVE 체크인 조회(없으면 null)",
     "높음", "예정", "1.0", "GET /api/check-ins/me"),
    ("CHECKIN-007", "체크인", "체크인", "전역 '혼밥 중' 상태 바",
     "앱 어디서나 현재 혼밥 상태 표시 + '끝내기'로 종료(클라 전역 상태)",
     "높음", "예정", "1.0", "06 §1.6"),
    ("CHECKIN-008", "체크인", "통계", "혼밥 유저 수 표시",
     "'오늘 혼밥 N명' / '현재 혼밥 중 N명' 집계 표시(사회적 증거)",
     "높음", "예정", "1.0", "FR-103 / GET /api/check-ins/stats"),
    ("CHECKIN-009", "체크인", "통계", "지도 기반 혼밥러 표시",
     "지도에서 식당별 현재 혼밥러 수를 핀/마커로 표시(반경 검색)",
     "높음", "예정", "1.0", "FR-104 / GET /api/check-ins/map"),
    ("CHECKIN-010", "체크인", "연결", "같은 식당 혼밥러 목록",
     "선택 식당의 현재 ACTIVE 혼밥러 목록(닉네임/체크인 경과만, 좌표·실명 비노출)",
     "높음", "예정", "1.0", "FR-107 / GET /api/places/{placeId}/check-ins"),

    # ───────── 같이먹기 ─────────
    ("MEAL-001", "같이먹기", "신청", "같이먹기 신청(인사말)",
     "같은 식당 혼밥러(대상 체크인)에게 인사말 포함 신청 전송",
     "높음", "예정", "1.0", "FR-108 / POST /api/meal-requests"),
    ("MEAL-002", "같이먹기", "신청", "중복·자기 신청 방지",
     "UNIQUE(from_user, to_check_in)로 중복 방지, 자기 자신 신청 차단",
     "높음", "예정", "1.0", "ERD C-2"),
    ("MEAL-003", "같이먹기", "신청", "수신 거부자 차단",
     "opt-in off(allow_meal_request=false) 대상에게 신청 불가(403)",
     "높음", "예정", "1.0", "FR-108 / NFR-03"),
    ("MEAL-004", "같이먹기", "조회", "받은 신청 목록",
     "받은 같이먹기 신청 목록 조회(상태 필터)",
     "높음", "예정", "1.0", "GET /api/meal-requests?role=received"),
    ("MEAL-005", "같이먹기", "조회", "보낸 신청 목록",
     "보낸 같이먹기 신청 목록 조회",
     "높음", "예정", "1.0", "GET /api/meal-requests?role=sent"),
    ("MEAL-006", "같이먹기", "응답", "신청 수락",
     "신청 수락 → status ACCEPTED + responded_at 기록",
     "높음", "예정", "1.0", "PATCH /api/meal-requests/{id}/accept"),
    ("MEAL-007", "같이먹기", "응답", "신청 거절",
     "신청 거절 → status DECLINED + responded_at 기록",
     "높음", "예정", "1.0", "PATCH /api/meal-requests/{id}/decline"),
    ("MEAL-008", "같이먹기", "알림", "신청 수신 알림",
     "같이먹기 신청 수신/수락 시 사용자 알림(푸시/폴링)",
     "높음", "예정", "1.0", "NFR-08"),

    # ───────── 메이트 ─────────
    ("MATE-001", "메이트", "신청", "메이트 신청",
     "다른 사용자에게 메이트 신청(PENDING). 이미 메이트/차단/자기자신 가드",
     "중간", "예정", "2.0", "FR-221 / POST /api/mate-requests"),
    ("MATE-002", "메이트", "응답", "메이트 신청 수락",
     "수락 시 mates에 양방향 2행 생성 + 신청 종료 이력 이관",
     "중간", "예정", "2.0", "FR-221"),
    ("MATE-003", "메이트", "응답", "메이트 신청 거절",
     "거절 시 신청 종료 → 이력 이관(mates 미생성)",
     "중간", "예정", "2.0", "FR-221"),
    ("MATE-004", "메이트", "이력", "메이트 신청 이력 보관/조회",
     "종료(수락/거절/취소) 신청을 mate_request_history로 이관·과거 이력 조회",
     "중간", "예정", "2.0", "ERD D-3"),
    ("MATE-005", "메이트", "신청", "재신청 허용",
     "진행 중(PENDING)만 보관하므로 거절/해제 후 같은 상대 재신청 가능",
     "중간", "예정", "2.0", "ERD D-3"),
    ("MATE-006", "메이트", "조회", "내 메이트 목록",
     "내 메이트 목록 조회(양방향 2행으로 단순 조회)",
     "중간", "예정", "2.0", "GET /api/mates"),
    ("MATE-007", "메이트", "조회", "메이트 온라인 상태",
     "메이트의 현재 ACTIVE 체크인(지금 혼밥 중)·장소 표시",
     "중간", "예정", "2.0", "ERD D-2"),
    ("MATE-008", "메이트", "조회", "메이트 프로필(타인)",
     "타인 메이트 프로필 조회(공개 즐겨찾기 등)",
     "중간", "예정", "2.0", "06 D-4"),
    ("MATE-009", "메이트", "관계", "메이트 해제",
     "메이트 관계 해제(양방향 2행 삭제)",
     "중간", "예정", "2.0", "DELETE /api/mates/{userId}"),
    ("MATE-010", "메이트", "추천", "추천 메이트",
     "함께 아는 사람 기반 메이트 추천",
     "낮음", "예정", "3.0", "FR-222 / GET /api/mates/suggestions"),
    ("MATE-011", "메이트", "통계", "'같이 N회' 집계",
     "두 사람 공통 식사 횟수 파생 집계(check_ins/reviews, 별도 컬럼 없음)",
     "중간", "예정", "2.0", "ERD D-2"),

    # ───────── 즐겨찾기 ─────────
    ("FAV-001", "즐겨찾기", "그룹", "즐겨찾기 그룹 생성",
     "이름·아이콘·설명으로 즐겨찾기 그룹 생성",
     "중간", "예정", "2.0", "FR-211 / favorite_groups"),
    ("FAV-002", "즐겨찾기", "그룹", "그룹 공개 설정",
     "그룹 공개 여부(is_public) 설정 — MateProfile 공개 노출 연동",
     "중간", "예정", "2.0", "FR-211"),
    ("FAV-003", "즐겨찾기", "식당", "그룹에 식당 담기",
     "그룹에 식당 추가(favorite_places, 그룹+식당 UNIQUE)",
     "중간", "예정", "2.0", "FR-211"),
    ("FAV-004", "즐겨찾기", "식당", "'다녀옴' 표시",
     "체크인/리뷰 연계로 '다녀옴' 산출·표시",
     "중간", "예정", "2.0", "FR-211"),
    ("FAV-005", "즐겨찾기", "조회", "즐겨찾기 목록/그룹 조회",
     "내 즐겨찾기 그룹·식당 목록 조회",
     "중간", "예정", "2.0", "06 E"),
    ("FAV-006", "즐겨찾기", "식당", "식당 즐겨찾기 추가",
     "식당 상세에서 즐겨찾기(♡) 추가",
     "중간", "예정", "2.0", "POST /api/favorites/places"),

    # ───────── 리뷰 · 혼밥일기(통합) ─────────
    ("REVIEW-001", "리뷰·혼밥일기", "작성", "리뷰=혼밥일기 작성",
     "한 번 작성으로 공개 리뷰 겸 개인 방문기록 저장(checkInId·placeId·visitedAt·mood·content). 좋아요/댓글 없음",
     "중간", "예정", "2.0", "FR-201E·FR-204 / POST /api/reviews"),
    ("REVIEW-002", "리뷰·혼밥일기", "작성", "별점 2종 입력",
     "맛 별점(taste_rating)·혼밥 친화 별점(solo_friendly_rating) 1~5 입력",
     "중간", "예정", "2.0", "FR-201E"),
    ("REVIEW-003", "리뷰·혼밥일기", "작성", "친화태그 입력",
     "혼밥 친화태그 입력(review_tags, place_id 역정규화로 식당별 집계)",
     "중간", "예정", "2.0", "ERD F-3"),
    ("REVIEW-004", "리뷰·혼밥일기", "작성", "리뷰 사진 업로드",
     "리뷰 사진 다중 업로드(review_photos) — 식당 사진 갤러리 출처",
     "중간", "예정", "2.0", "ERD F-2"),
    ("REVIEW-005", "리뷰·혼밥일기", "정책", "같은 식당 다회 작성",
     "같은 식당에 방문마다 작성 허용(UNIQUE 미적용)",
     "중간", "예정", "2.0", "ERD §7 #3"),
    ("REVIEW-006", "리뷰·혼밥일기", "조회", "식당 리뷰 목록",
     "식당별 리뷰(=혼밥일기) 목록 조회(별점2·사진·친화태그)",
     "중간", "예정", "2.0", "GET /api/places/{placeId}/reviews"),
    ("REVIEW-007", "리뷰·혼밥일기", "조회", "내 혼밥 기록(월별)",
     "내 방문+리뷰(=일기) 월별 기록 조회",
     "중간", "예정", "2.0", "FR-203E / GET /api/me/reviews"),

    # ───────── 활동기록 · 통계 ─────────
    ("ACT-001", "활동기록·통계", "통계", "활동 요약 통계",
     "총 혼밥/리뷰/방문 식당/이번 달 등 개인 활동 요약 카운트",
     "중간", "예정", "2.0", "FR-203E / GET /api/me/dining-stats"),

    # ───────── 성취(뱃지·챌린지) ─────────
    ("ACHV-001", "성취", "뱃지", "뱃지 정의(마스터)",
     "뱃지 코드·이름·이모지·조건 정의(badges)",
     "중간", "예정", "2.0", "FR-202E / 레벨은 보류(C9)"),
    ("ACHV-002", "성취", "뱃지", "뱃지 자동 부여",
     "체크인/일기/메이트 트리거로 시스템 자동 부여(user_badges)",
     "중간", "예정", "2.0", "FR-202E"),
    ("ACHV-003", "성취", "뱃지", "내 뱃지 조회",
     "획득/잠금 뱃지 목록 조회",
     "중간", "예정", "2.0", "FR-202E"),
    ("ACHV-004", "성취", "챌린지", "챌린지(주간 미션)",
     "주간/월간 챌린지 정의·참여·진행률(challenges/user_challenges)",
     "낮음", "예정", "3.0", "FR-301E"),

    # ───────── 알림 ─────────
    ("NOTI-001", "알림", "푸시", "디바이스 토큰 등록",
     "푸시 디바이스 토큰 등록/갱신(device_tokens, Expo/FCM)",
     "높음", "예정", "1.0", "FR-232 / NFR-08"),
    ("NOTI-002", "알림", "푸시", "푸시 알림 발송",
     "같이먹기 신청/수락, 메이트 신청/혼밥 시작 등 트리거 기반 푸시 발송",
     "높음", "예정", "1.0", "FR-232 / NFR-08"),
    ("NOTI-003", "알림", "인앱", "인앱 알림함",
     "인앱 알림 목록·읽음 처리(notifications)",
     "중간", "예정", "2.0", "FR-232"),
    ("NOTI-004", "알림", "설정", "알림 설정(카테고리)",
     "전체/활동/메이트/마케팅 카테고리별 토글(notification_settings)",
     "중간", "예정", "2.0", "FR-231"),
    ("NOTI-005", "알림", "설정", "방해 금지 시간",
     "방해 금지(DND) 시작/종료 시간 설정",
     "중간", "예정", "2.0", "FR-231"),

    # ───────── 안전(차단·신고) ─────────
    ("SAFE-001", "안전", "차단", "사용자 차단",
     "사용자 차단(blocks, 중복 방지)",
     "중간", "예정", "2.0", "FR-242 / NFR-03"),
    ("SAFE-002", "안전", "신고", "신고",
     "사용자/리뷰/체크인/댓글 대상 신고 접수(reports, 다형 참조)",
     "중간", "예정", "2.0", "FR-242"),
    ("SAFE-003", "안전", "차단", "차단 연동 로직",
     "차단 시 상호 노출 차단 + 같이먹기/메이트 차단(서비스 레이어)",
     "중간", "예정", "2.0", "ERD §7 #5"),

    # ───────── 운영(공지·FAQ·문의) ─────────
    ("OPS-001", "운영", "공지", "공지사항",
     "관리자 공지사항 작성/노출(notices, 핀·게시일)",
     "낮음", "예정", "3.0", "FR-241"),
    ("OPS-002", "운영", "FAQ", "자주 묻는 질문",
     "FAQ 목록/검색(faqs)",
     "낮음", "예정", "3.0", "FR-243"),
    ("OPS-003", "운영", "문의", "1:1 문의",
     "고객 1:1 문의 접수/답변(inquiries, OPEN/ANSWERED/CLOSED)",
     "낮음", "예정", "3.0", "FR-243"),

    # ───────── 공통 · 시스템 ─────────
    ("COMMON-001", "공통·시스템", "인증/인가", "JWT 인증/인가",
     "JWT 기반 인증(Access/Refresh 분리), 인증 필요 API 토큰 검증(401)",
     "높음", "예정", "1.0", "NFR-01"),
    ("COMMON-002", "공통·시스템", "파일", "이미지 업로드",
     "프로필/리뷰 사진 업로드(오브젝트 스토리지 S3 등)",
     "높음", "예정", "1.0", "POST /api/files"),
    ("COMMON-003", "공통·시스템", "프라이버시", "위치정보 최소수집",
     "위치정보 목적 최소 수집, 정확 좌표 대신 식당/구역 단위 표시, 실명 비노출",
     "높음", "예정", "1.0", "NFR-03"),
    ("COMMON-004", "공통·시스템", "API 규약", "공통 응답/에러 포맷",
     "success/data·error(code·message) 공통 포맷, 표준 상태 코드(400/401/403/404/409)",
     "높음", "예정", "1.0", "05 §1"),
    ("COMMON-005", "공통·시스템", "API 규약", "페이징 규약",
     "page/size(0-base) + content·totalElements 응답 규약",
     "중간", "예정", "1.0", "05 §1.4"),

    # ───────── 보류 ─────────
    ("REACT-001", "연결(보류)", "반응", "초간단 반응",
     "'나도 여기 있음'/'같이 먹는 중' 초간단 반응 — 이번 범위 보류(추후 도입). 핵심 연결은 같이먹기로 성립",
     "낮음", "보류", "-", "FR-106 / 보류"),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "기능요구사항"

    # 스타일
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="맑은 고딕", size=10)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 헤더
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 데이터
    imp_fill = {
        "높음": PatternFill("solid", fgColor="FCE4E4"),
        "중간": PatternFill("solid", fgColor="FFF6E0"),
        "낮음": PatternFill("solid", fgColor="EAF3FB"),
    }
    state_fill = {"보류": PatternFill("solid", fgColor="EEEEEE")}

    for r, row in enumerate(ROWS, start=2):
        ws.append(row)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = cell_font
            cell.border = border
            # 중앙정렬 컬럼: ID(1), 중요도(6), 상태(7), 버전(8)
            if c in (1, 6, 7, 8):
                cell.alignment = center
            else:
                cell.alignment = wrap_top
        # 중요도 색
        ws.cell(row=r, column=6).fill = imp_fill.get(row[5], PatternFill())
        # 보류 상태 회색
        if row[6] in state_fill:
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = state_fill["보류"]

    # 열 너비
    widths = [12, 14, 12, 20, 60, 8, 8, 7, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 틀 고정 + 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS)+1}"

    out = "/Users/yoonhyunwoo/project/docs/혼정-기능요구사항-목록.xlsx"
    wb.save(out)
    print(f"saved: {out}")
    print(f"rows: {len(ROWS)}")
    # 대분류별 카운트
    from collections import Counter
    cnt = Counter(r[1] for r in ROWS)
    for k, v in cnt.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    build()
